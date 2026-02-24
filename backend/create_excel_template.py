"""
Create the Accounting Excel Template with sample data structure.
This script generates the template that users will fill out for import.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from decimal import Decimal
import os


def create_excel_template():
    """Create the full Accounting Excel Template."""
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create all required sheets
    create_business_config_sheet(wb)
    create_accounts_sheet(wb)
    create_categories_sheet(wb)
    create_tax_rates_sheet(wb)
    
    # Create transaction sheets for each month
    for month in range(1, 13):
        create_transactions_sheet(wb, month)
    
    # Save the workbook
    output_path = "/home/skai8888/code/other projects/Accounting tool/10-mvp/00-source-excel/Accounting-Excel-Template.xlsx"
    wb.save(output_path)
    print(f"Created Excel template at: {output_path}")
    return output_path


def style_header(cell):
    """Apply header styling."""
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")


def style_subheader(cell):
    """Apply subheader styling."""
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")


def create_business_config_sheet(wb):
    """Create Business Configuration sheet."""
    ws = wb.create_sheet("Business Config", 0)
    
    ws['A1'] = "Business Configuration"
    ws['A1'].font = Font(bold=True, size=14)
    
    # Config fields
    config = [
        ("Business Name", "My Company Ltd", "Required - Your business name"),
        ("Fiscal Year Start Month", 1, "1-12 (January=1, December=12)"),
        ("Currency", "CHF", "ISO currency code (e.g., CHF, USD, EUR)"),
        ("Tax Authority", "Swiss VAT", "Name of tax authority"),
    ]
    
    ws['A3'] = "Field"
    ws['B3'] = "Value"
    ws['C3'] = "Description"
    
    for col in ['A', 'B', 'C']:
        style_header(ws[f'{col}3'])
    
    row = 4
    for field, value, desc in config:
        ws[f'A{row}'] = field
        ws[f'B{row}'] = value
        ws[f'C{row}'] = desc
        row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 50


def create_accounts_sheet(wb):
    """Create Accounts sheet."""
    ws = wb.create_sheet("Accounts")
    
    ws['A1'] = "Accounts (Bank, Credit Card, Assets)"
    ws['A1'].font = Font(bold=True, size=14)
    
    # Headers
    headers = ["Account Name", "Type", "Opening Balance", "Currency", "Notes"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        style_header(cell)
    
    # Sample data
    accounts = [
        ("Bank Account #1", "bank", 10000.00, "CHF", "Primary checking"),
        ("Bank Account #2", "bank", 5000.00, "CHF", "Savings"),
        ("Credit Card Account", "credit_card", 0.00, "CHF", "Company credit card"),
    ]
    
    row = 4
    for name, acc_type, balance, currency, notes in accounts:
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=acc_type)
        ws.cell(row=row, column=3, value=balance)
        ws.cell(row=row, column=4, value=currency)
        ws.cell(row=row, column=5, value=notes)
        row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 30


def create_categories_sheet(wb):
    """Create Categories sheet (Chart of Accounts)."""
    ws = wb.create_sheet("Categories")
    
    ws['A1'] = "Chart of Accounts Categories"
    ws['A1'].font = Font(bold=True, size=14)
    
    # Headers
    headers = ["Code", "Name", "Type", "Report"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        style_header(cell)
    
    # Categories (26 total: 5 income, 6 COGS, 15 expense)
    categories = []
    
    # Income (head_1 to head_5)
    for i in range(1, 6):
        categories.append((f"head_{i}", f"Income Category {i}", "income", "pl"))
    
    # COGS (head_6 to head_11)
    for i in range(6, 12):
        categories.append((f"head_{i}", f"COGS Category {i-5}", "cogs", "pl"))
    
    # Expenses (head_12 to head_26)
    for i in range(12, 27):
        categories.append((f"head_{i}", f"Expense Category {i-11}", "expense", "pl"))
    
    row = 4
    for code, name, cat_type, report in categories:
        ws.cell(row=row, column=1, value=code)
        ws.cell(row=row, column=2, value=name)
        ws.cell(row=row, column=3, value=cat_type)
        ws.cell(row=row, column=4, value=report)
        row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10


def create_tax_rates_sheet(wb):
    """Create Tax Rates sheet."""
    ws = wb.create_sheet("Tax Rates")
    
    ws['A1'] = "Tax Rates (VAT/Sales Tax)"
    ws['A1'].font = Font(bold=True, size=14)
    
    # Headers
    headers = ["Name", "Rate (decimal)", "Description"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        style_header(cell)
    
    # Sample tax rates
    tax_rates = [
        ("VAT Exempt", 0.0, "No VAT applied"),
        ("VAT 2.5%", 0.025, "Reduced VAT rate"),
        ("VAT 8.1%", 0.081, "Standard Swiss VAT rate"),
    ]
    
    row = 4
    for name, rate, desc in tax_rates:
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=rate)
        ws.cell(row=row, column=3, value=desc)
        row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 35


def create_transactions_sheet(wb, month):
    """Create a transaction sheet for a specific month."""
    ws = wb.create_sheet(f"Month{month}")
    
    ws['A1'] = f"Transactions - Month {month}"
    ws['A1'].font = Font(bold=True, size=14)
    
    # Headers
    headers = [
        "Date (YYYY-MM-DD)",
        "Account Name",
        "Payee",
        "Description",
        "Reference",
        "Direction (in/out)",
        "Gross Amount",
        "Tax Rate Name",
        "Category Code(s)",
        "Special Type",
        "Allocation Amount(s)",
        "Is Reconciled (yes/no)"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        style_header(cell)
    
    # Sample data row with instructions
    sample = [
        f"2024-{month:02d}-15",  # Date
        "Bank Account #1",       # Account
        "Sample Vendor",         # Payee
        "Sample transaction",    # Description
        "INV-001",               # Reference
        "out",                   # Direction
        1000.00,                 # Gross Amount
        "VAT 8.1%",              # Tax Rate
        "head_12",               # Category
        "",                      # Special Type
        1000.00,                 # Allocation
        "no"                     # Reconciled
    ]
    
    for col, value in enumerate(sample, 1):
        cell = ws.cell(row=4, column=col, value=value)
        cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    # Notes
    ws['A6'] = "Notes:"
    ws['A6'].font = Font(bold=True)
    ws['A7'] = "- For multiple categories, separate codes with semicolons (e.g., 'head_1;head_2')"
    ws['A8'] = "- For multiple allocations, separate amounts with semicolons (must match number of categories)"
    ws['A9'] = "- Special types: capital, loan_in, loan_repayment, transfer_in, transfer_out, asset_purchase, tax_payment, drawings, income_tax, payroll_tax"
    ws['A10'] = "- Direction: 'in' for income/money received, 'out' for expense/money paid"
    
    # Adjust column widths
    widths = [18, 20, 20, 30, 15, 18, 15, 18, 20, 15, 22, 22]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width


if __name__ == "__main__":
    create_excel_template()
