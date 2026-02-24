"""
Excel Import Service - Import business data from Excel template.

Parses the Accounting-Excel-Template.xlsx file and imports:
- Business configuration
- Accounts
- Categories
- Tax Rates
- Transactions (Month1-12)
"""
from datetime import datetime, date
from decimal import Decimal, InvalidOperation
from typing import Dict, List, Optional, Tuple, Any
import io

from openpyxl import load_workbook
from sqlalchemy.orm import Session
from fastapi import UploadFile

from . import models, schemas, crud


class ExcelImportError(Exception):
    """Custom exception for Excel import errors."""
    pass


class ExcelImportService:
    """Service for importing accounting data from Excel files."""
    
    def __init__(self, db: Session):
        self.db = db
        self.errors: List[str] = []
        self.warnings: List[str] = []
    
    def import_excel(self, file: UploadFile, business_id: Optional[int] = None) -> Dict[str, Any]:
        """
        Import data from an Excel file.
        
        Args:
            file: Uploaded Excel file
            business_id: Optional existing business ID to update (creates new if None)
        
        Returns:
            Dictionary with import results
        """
        self.errors = []
        self.warnings = []
        
        try:
            # Read the uploaded file
            contents = file.file.read()
            workbook = load_workbook(io.BytesIO(contents))
        except Exception as e:
            raise ExcelImportError(f"Failed to read Excel file: {str(e)}")
        
        # Get sheet names
        sheet_names = workbook.sheetnames
        
        result = {
            "success": False,
            "business_id": None,
            "business_name": None,
            "accounts_imported": 0,
            "categories_imported": 0,
            "tax_rates_imported": 0,
            "transactions_imported": 0,
            "errors": [],
            "warnings": [],
        }
        
        try:
            # Import business config
            business = self._import_business_config(workbook, business_id)
            result["business_id"] = business.id
            result["business_name"] = business.name
            
            # Import accounts
            accounts = self._import_accounts(workbook, business.id)
            result["accounts_imported"] = len(accounts)
            
            # Import categories
            categories = self._import_categories(workbook, business.id)
            result["categories_imported"] = len(categories)
            
            # Import tax rates
            tax_rates = self._import_tax_rates(workbook, business.id)
            result["tax_rates_imported"] = len(tax_rates)
            
            # Import transactions from all month sheets
            total_txns = 0
            for month in range(1, 13):
                sheet_name = f"Month{month}"
                if sheet_name in sheet_names:
                    txns = self._import_transactions(workbook, sheet_name, business.id, month)
                    total_txns += len(txns)
            
            result["transactions_imported"] = total_txns
            result["success"] = len(self.errors) == 0
            
        except ExcelImportError as e:
            self.errors.append(str(e))
        
        result["errors"] = self.errors
        result["warnings"] = self.warnings
        
        return result
    
    def _import_business_config(self, workbook, existing_business_id: Optional[int]) -> models.Business:
        """Import business configuration from Excel."""
        if "Business Config" not in workbook.sheetnames:
            raise ExcelImportError("Missing 'Business Config' sheet")
        
        ws = workbook["Business Config"]
        
        # Parse config values
        config = {}
        for row in range(4, 20):  # Read rows 4-19
            field_cell = ws.cell(row=row, column=1).value
            value_cell = ws.cell(row=row, column=2).value
            
            if field_cell and value_cell:
                config[field_cell.strip()] = value_cell
        
        # Extract values
        name = config.get("Business Name", "Imported Business")
        fiscal_month = self._parse_int(config.get("Fiscal Year Start Month"), 1)
        currency = config.get("Currency", "CHF")
        
        # Validate fiscal month
        if not (1 <= fiscal_month <= 12):
            self.warnings.append(f"Invalid fiscal year start month ({fiscal_month}), using 1 (January)")
            fiscal_month = 1
        
        # Create or update business
        if existing_business_id:
            business = crud.get_business(self.db, existing_business_id)
            if not business:
                raise ExcelImportError(f"Business with ID {existing_business_id} not found")
            
            # Update business
            business.name = name
            business.fiscal_year_start_month = fiscal_month
            business.currency = currency
            self.db.commit()
            self.db.refresh(business)
        else:
            # Create new business
            business_data = schemas.BusinessCreate(
                name=name,
                fiscal_year_start_month=fiscal_month,
                currency=currency,
            )
            business = crud.create_business(self.db, business_data)
            # Clear default categories and accounts (we'll import our own)
            self.db.query(models.Category).filter(models.Category.business_id == business.id).delete()
            self.db.query(models.Account).filter(models.Account.business_id == business.id).delete()
            self.db.commit()
        
        return business
    
    def _import_accounts(self, workbook, business_id: int) -> List[models.Account]:
        """Import accounts from Excel."""
        if "Accounts" not in workbook.sheetnames:
            self.warnings.append("Missing 'Accounts' sheet, skipping account import")
            return []
        
        ws = workbook["Accounts"]
        accounts = []
        
        # Read from row 4 onwards (skip headers)
        row = 4
        while True:
            name = ws.cell(row=row, column=1).value
            if not name:
                break
            
            acc_type = ws.cell(row=row, column=2).value or "bank"
            opening_balance = self._parse_decimal(ws.cell(row=row, column=3).value, Decimal("0.00"))
            
            # Validate account type
            valid_types = ["bank", "credit_card", "asset"]
            if acc_type not in valid_types:
                self.warnings.append(f"Row {row}: Invalid account type '{acc_type}', using 'bank'")
                acc_type = "bank"
            
            # Create account
            account_data = schemas.AccountCreate(
                name=name,
                type=acc_type,
                opening_balance=opening_balance,
            )
            account = crud.create_account(self.db, business_id, account_data)
            accounts.append(account)
            
            row += 1
            
            # Safety limit
            if row > 1000:
                self.warnings.append("Account import stopped at 1000 accounts")
                break
        
        return accounts
    
    def _import_categories(self, workbook, business_id: int) -> List[models.Category]:
        """Import categories from Excel."""
        if "Categories" not in workbook.sheetnames:
            self.warnings.append("Missing 'Categories' sheet, skipping category import")
            return []
        
        ws = workbook["Categories"]
        categories = []
        
        # Read from row 4 onwards
        row = 4
        while True:
            code = ws.cell(row=row, column=1).value
            if not code:
                break
            
            name = ws.cell(row=row, column=2).value or f"Category {code}"
            cat_type = ws.cell(row=row, column=3).value or "expense"
            report = ws.cell(row=row, column=4).value or "pl"
            
            # Validate type
            valid_types = ["income", "cogs", "expense"]
            if cat_type not in valid_types:
                self.warnings.append(f"Row {row}: Invalid category type '{cat_type}', using 'expense'")
                cat_type = "expense"
            
            # Validate report
            valid_reports = ["pl", "bs"]
            if report not in valid_reports:
                report = "pl"
            
            # Create category
            category_data = schemas.CategoryCreate(
                code=code,
                name=name,
                type=cat_type,
                report=report,
            )
            category = crud.create_category(self.db, business_id, category_data)
            categories.append(category)
            
            row += 1
            
            if row > 1000:
                self.warnings.append("Category import stopped at 1000 categories")
                break
        
        return categories
    
    def _import_tax_rates(self, workbook, business_id: int) -> List[models.TaxRate]:
        """Import tax rates from Excel."""
        if "Tax Rates" not in workbook.sheetnames:
            self.warnings.append("Missing 'Tax Rates' sheet, skipping tax rate import")
            return []
        
        ws = workbook["Tax Rates"]
        tax_rates = []
        
        # Read from row 4 onwards
        row = 4
        while True:
            name = ws.cell(row=row, column=1).value
            if not name:
                break
            
            rate_value = ws.cell(row=row, column=2).value
            rate = self._parse_decimal(rate_value, Decimal("0.00"))
            
            # Validate rate is between 0 and 1
            if rate < 0 or rate >= 1:
                self.warnings.append(f"Row {row}: Invalid tax rate {rate}, using 0.00")
                rate = Decimal("0.00")
            
            # Create tax rate
            tax_rate_data = schemas.TaxRateCreate(
                name=name,
                rate=rate,
            )
            tax_rate = crud.create_tax_rate(self.db, business_id, tax_rate_data)
            tax_rates.append(tax_rate)
            
            row += 1
            
            if row > 1000:
                self.warnings.append("Tax rate import stopped at 1000 rates")
                break
        
        return tax_rates
    
    def _import_transactions(self, workbook, sheet_name: str, business_id: int, month: int) -> List[models.Transaction]:
        """Import transactions from a month sheet."""
        ws = workbook[sheet_name]
        transactions = []
        
        # Get account mapping
        accounts = crud.get_accounts_by_business(self.db, business_id)
        account_map = {a.name: a for a in accounts}
        
        # Get category mapping
        categories = crud.get_categories_by_business(self.db, business_id)
        category_map = {c.code: c for c in categories}
        
        # Get tax rate mapping
        tax_rates = crud.get_tax_rates_by_business(self.db, business_id)
        tax_rate_map = {t.name: t for t in tax_rates}
        
        # Read from row 4 onwards
        row = 4
        while True:
            date_cell = ws.cell(row=row, column=1).value
            if not date_cell:
                # Check if this is an empty row or end of data
                # Look at next few cells to confirm
                has_data = False
                for col in range(2, 10):
                    if ws.cell(row=row, column=col).value:
                        has_data = True
                        break
                if not has_data:
                    break
            
            try:
                # Parse date
                txn_date = self._parse_date(date_cell)
                if not txn_date:
                    self.warnings.append(f"{sheet_name} Row {row}: Invalid date '{date_cell}', skipping")
                    row += 1
                    continue
                
                # Parse account
                account_name = ws.cell(row=row, column=2).value
                if not account_name or account_name not in account_map:
                    self.warnings.append(f"{sheet_name} Row {row}: Unknown account '{account_name}', skipping")
                    row += 1
                    continue
                
                account = account_map[account_name]
                
                # Parse other fields
                payee = ws.cell(row=row, column=3).value or ""
                description = ws.cell(row=row, column=4).value or ""
                reference = ws.cell(row=row, column=5).value or ""
                direction = ws.cell(row=row, column=6).value or "out"
                gross_amount = self._parse_decimal(ws.cell(row=row, column=7).value, Decimal("0.00"))
                
                if gross_amount <= 0:
                    self.warnings.append(f"{sheet_name} Row {row}: Amount must be positive, skipping")
                    row += 1
                    continue
                
                # Parse tax rate
                tax_rate_name = ws.cell(row=row, column=8).value
                tax_rate_id = None
                if tax_rate_name and tax_rate_name in tax_rate_map:
                    tax_rate_id = tax_rate_map[tax_rate_name].id
                
                # Parse categories and allocations
                category_codes = ws.cell(row=row, column=9).value
                special_type = ws.cell(row=row, column=10).value
                allocation_amounts = ws.cell(row=row, column=11).value
                
                # Parse reconciled status
                reconciled = ws.cell(row=row, column=12).value
                is_reconciled = str(reconciled).lower() in ["yes", "true", "1"]
                
                # Build allocations
                allocations = []
                
                # Handle category-based allocations
                if category_codes:
                    codes = [c.strip() for c in str(category_codes).split(";") if c.strip()]
                    amounts = []
                    
                    if allocation_amounts:
                        amounts = [self._parse_decimal(a.strip(), Decimal("0.00")) 
                                  for a in str(allocation_amounts).split(";") if a.strip()]
                    
                    # If no amounts specified, distribute equally
                    if not amounts:
                        amount_per = gross_amount / len(codes)
                        amounts = [amount_per] * len(codes)
                    
                    # Match codes with amounts
                    for i, code in enumerate(codes):
                        if code in category_map:
                            amount = amounts[i] if i < len(amounts) else amounts[-1]
                            allocations.append(schemas.TransactionAllocation(
                                category_id=category_map[code].id,
                                amount=amount
                            ))
                        else:
                            self.warnings.append(f"{sheet_name} Row {row}: Unknown category '{code}'")
                
                # Handle special type allocation
                if special_type:
                    valid_special_types = [
                        "capital", "loan_in", "loan_repayment", "transfer_in",
                        "transfer_out", "asset_purchase", "tax_payment",
                        "drawings", "income_tax", "payroll_tax"
                    ]
                    if special_type in valid_special_types:
                        allocations.append(schemas.TransactionAllocation(
                            special_type=special_type,
                            amount=gross_amount
                        ))
                    else:
                        self.warnings.append(f"{sheet_name} Row {row}: Unknown special type '{special_type}'")
                
                # If no allocations, create one with uncategorized (use first expense category)
                if not allocations:
                    expense_cats = [c for c in categories if c.type == models.CategoryType.EXPENSE]
                    if expense_cats:
                        allocations.append(schemas.TransactionAllocation(
                            category_id=expense_cats[0].id,
                            amount=gross_amount
                        ))
                    else:
                        self.warnings.append(f"{sheet_name} Row {row}: No allocations created, skipping")
                        row += 1
                        continue
                
                # Create transaction
                txn_data = schemas.TransactionCreate(
                    date=txn_date,
                    payee=payee,
                    description=description,
                    reference=reference,
                    direction=direction,
                    gross_amount=gross_amount,
                    tax_rate_id=tax_rate_id,
                    allocations=allocations,
                )
                
                txn = crud.create_transaction(self.db, account.id, txn_data)
                
                # Set reconciled status if needed
                if is_reconciled:
                    txn.is_reconciled = True
                    self.db.commit()
                
                transactions.append(txn)
                
            except Exception as e:
                self.warnings.append(f"{sheet_name} Row {row}: Error importing - {str(e)}")
            
            row += 1
            
            if row > 10000:
                self.warnings.append(f"{sheet_name} import stopped at 10000 rows")
                break
        
        return transactions
    
    def _parse_int(self, value: Any, default: int) -> int:
        """Parse integer from various formats."""
        if value is None:
            return default
        try:
            return int(value)
        except (ValueError, TypeError):
            return default
    
    def _parse_decimal(self, value: Any, default: Decimal) -> Decimal:
        """Parse decimal from various formats."""
        if value is None:
            return default
        try:
            if isinstance(value, str):
                # Handle common formats
                value = value.replace(",", "").replace("'", "")
            return Decimal(str(value))
        except (ValueError, TypeError, InvalidOperation):
            return default
    
    def _parse_date(self, value: Any) -> Optional[date]:
        """Parse date from various formats."""
        if value is None:
            return None
        
        try:
            if isinstance(value, datetime):
                return value.date()
            if isinstance(value, date):
                return value
            if isinstance(value, str):
                # Try ISO format (YYYY-MM-DD)
                try:
                    return datetime.strptime(value.strip(), "%Y-%m-%d").date()
                except ValueError:
                    pass
                # Try European format (DD.MM.YYYY)
                try:
                    return datetime.strptime(value.strip(), "%d.%m.%Y").date()
                except ValueError:
                    pass
                # Try US format (MM/DD/YYYY)
                try:
                    return datetime.strptime(value.strip(), "%m/%d/%Y").date()
                except ValueError:
                    pass
            return None
        except (ValueError, TypeError):
            return None
